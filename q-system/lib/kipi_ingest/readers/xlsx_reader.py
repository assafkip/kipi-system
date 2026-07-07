"""XLSX reader. Fixes the LIVE analog losses the audit found:
  - first-sheet-only: reads EVERY sheet (page = sheet ordinal), not just the first.
  - formula-cell loss: `data_only=True` yields None for an uncomputed formula; this
    recovers the formula STRING so the cell is preserved, not silently blanked.
  - header assumption: every non-empty row (including the header) is captured, so a
    headerless or single-row sheet does not vanish.
  - silent sheet/row caps are COUNTED into the receipt as `truncated`.

openpyxl is imported lazily so importing this package never requires it.
"""
from __future__ import annotations

from pathlib import Path

from ..contract import Block, Drop, ReadResult, make_block_id


def read_xlsx(
    path: str | Path,
    *,
    max_sheets: int | None = None,
    max_rows_per_sheet: int | None = None,
) -> tuple[list[Block], ReadResult]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - environment-dependent
        raise RuntimeError("read_xlsx needs openpyxl (pip install openpyxl)") from e

    p = Path(path)
    values = load_workbook(p, read_only=True, data_only=True)
    formulas = load_workbook(p, read_only=True, data_only=False)
    try:
        sheet_names = values.sheetnames
        sheets_truncated = max_sheets is not None and len(sheet_names) > max_sheets
        use_sheets = sheet_names[:max_sheets] if sheets_truncated else sheet_names

        blocks: list[Block] = []
        dropped: list[Drop] = []
        attempted_rows = 0

        for si, name in enumerate(use_sheets, start=1):
            vws = values[name]
            fws = formulas[name]
            row_stream = zip(vws.iter_rows(values_only=True),
                             fws.iter_rows(values_only=True))
            emitted_here = 0
            for ri, (vrow, frow) in enumerate(row_stream, start=1):
                cells = []
                for v, f in zip(vrow, frow):
                    if v is None or v == "":
                        # formula-cell recovery: an uncomputed formula would be
                        # None here; keep its formula string rather than lose it.
                        if isinstance(f, str) and f.startswith("="):
                            cells.append(f)
                    else:
                        cells.append(str(v))
                if not cells:
                    continue
                attempted_rows += 1
                if max_rows_per_sheet is not None and emitted_here >= max_rows_per_sheet:
                    dropped.append(Drop(f"r.p{si}.{ri}",
                                        f"row cap {max_rows_per_sheet} on sheet {name!r}"))
                    continue
                blocks.append(Block(block_id=make_block_id("r", si, ri),
                                    unit="row", page=si, text=" | ".join(cells)))
                emitted_here += 1

        if sheets_truncated:
            for name in sheet_names[max_sheets:]:
                dropped.append(Drop(f"sheet:{name}",
                                    f"sheet cap {max_sheets}; sheet {name!r} not read"))

        return blocks, ReadResult(
            unit="row",
            attempted=attempted_rows,
            captured=len(blocks),
            truncated=bool(dropped),
            dropped=dropped,
        )
    finally:
        values.close()
        formulas.close()
